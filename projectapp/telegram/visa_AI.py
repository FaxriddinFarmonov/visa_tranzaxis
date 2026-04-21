


import re
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import FSInputFile
import asyncio
import os

TOKEN = "8556733636:AAG5uynF29fvJP1MR6E1ckJn9r5ZWyCMojE"

bot = Bot(token=TOKEN)
dp = Dispatcher()

# 📥 PDF
def extract_text_from_pdf(path):
    text = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text += t + "\n"
    return text


# 🔥 SMART SPLIT (yaxshilangan)
def smart_split(line):
    parts = re.split(r"\s{2,}", line.strip())

    if len(parts) == 1:
        parts = line.split()

    return parts


# 🔥 SUPER PARSER
def parse_rows(text):
    lines = text.split("\n")
    rows = []

    for line in lines:
        line = line.strip()
        if not line:
            rows.append([])
            continue

        # 🔹 Reporting (yonma-yon)
        if "Reporting for:" in line and "Settlement Currency Code:" in line:
            m = re.search(r"Reporting for:\s*(.*?)\s*Settlement Currency Code:\s*(\w+)", line)
            if m:
                rows.append(["Reporting for", m.group(1), "Currency", m.group(2)])
                continue

        # 🔹 Date line
        if "Settl. Date:" in line and "Clearing Currency Code:" in line:
            m = re.search(r"Settl\. Date:\s*(.*?)\s*Clearing Currency Code:\s*(\w+)", line)
            if m:
                rows.append(["Settl Date", m.group(1), "Clearing Currency", m.group(2)])
                continue

        # 🔹 TABLE HEADER
        if "Count" in line and "Interch." in line:
            rows.append(["Type", "Count", "Interch. value", "Reimb. Fees", "Net value"])
            continue

        # 🔹 ISSUER HEADERS
        if any(k in line for k in [
            "Issuer other",
            "Issuer originals",
            "Issuer reversals",
            "Issuer Total"
        ]):
            rows.append([line])
            continue

        # 🔹 OPTIONAL FEES HEADER
        if "Optional Issuer conversion fees" in line:
            rows.append(["Optional Fees", "Interch. amount", "Convers. Fee", "Opt. issuer fee"])
            continue

        # 🔹 VISA HEADER
        if "Visa Charges (Issuer)" in line:
            rows.append(["Visa Charges", "Amount"])
            continue

        # 🔹 VISA ROW (fix split)
        if "ISA CHARGE" in line:
            value = line.split()[-1]
            name = " ".join(line.split()[:-1])
            rows.append([name, value])
            continue

        # 🔹 TOTAL VISA
        if "Total for VISA charges" in line:
            rows.append(["Total VISA charges", line.split()[-1]])
            continue

        # 🔹 NET
        if "Net Settlement Amount" in line:
            rows.append(["Net Settlement", line.split()[-1]])
            continue

        # 🔹 OPTIONAL rows
        if line.startswith("Originals") or line.startswith("Total"):
            parts = smart_split(line)
            rows.append(parts)
            continue

        # 🔹 TABLE ROWS
        parts = smart_split(line)

        if len(parts) >= 4:
            name = " ".join(parts[:-4])
            nums = parts[-4:]
            rows.append([name] + nums)
        else:
            rows.append(parts)

    return rows


# 🎨 EXCEL DESIGN
def save_excel(rows, filename="output.xlsx"):
    wb = Workbook()
    ws = wb.active

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    fill_header = PatternFill(start_color="D9E1F2", fill_type="solid")

    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)

            # HEADER STYLE
            if val in ["Type", "Count", "Interch. value", "Reimb. Fees", "Net value"]:
                cell.font = bold
                cell.alignment = center
                cell.fill = fill_header
                cell.border = border

            # SECTION HEADER
            elif isinstance(val, str) and "Issuer" in val:
                cell.font = Font(bold=True, size=12)

            # IMPORTANT VALUES
            elif isinstance(val, str) and any(k in val for k in ["Total", "Settlement"]):
                cell.font = bold

            # TABLE CELLS
            if c_idx > 1:
                cell.border = border

    # AUTO WIDTH
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(filename)


# 🚀 BOT
@dp.message(Command("start"))
async def start(message: types.Message):
    await message.reply("PDF yubor — men uni PROFESSIONAL Excel qilaman 📊🔥")
# bnnbvnvbnvbv

@dp.message(F.document)
async def handle_file(message: types.Message):
    if not message.document.file_name.endswith(".pdf"):
        await message.reply("❌ PDF yubor")
        return

    os.makedirs("downloads", exist_ok=True)

    file = await bot.get_file(message.document.file_id)
    path = f"downloads/{message.document.file_name}"

    await bot.download_file(file.file_path, path)

    await message.reply("⏳ Processing...")

    text = extract_text_from_pdf(path)
    rows = parse_rows(text)

    save_excel(rows)

    await message.reply_document(FSInputFile("output.xlsx"))
    try:
        os.remove(path)
    except:
        pass

    # 🔥 Excelni ham o‘chirish
    try:
        os.remove("output.xlsx")
    except:
        pass

@dp.message()
async def other(message: types.Message):
    await message.reply("❌ Faqat PDF yubor")


async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())